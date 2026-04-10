"""
extract.py — 从 Excel 原始数据提取最高等级角色/武器属性
运行: python extract.py
输出: extracted_characters.json, extracted_weapons.json
"""
import json
import os
import glob
import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

DATA_DIR = os.path.dirname(os.path.abspath(__file__))


def read_sheet_as_dicts(ws):
    """将工作表转为 [{header: value, ...}, ...] 格式"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


def summarize_workbook(path):
    """打印工作簿所有 sheet 名和前3行，用于结构确认"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n{'='*60}")
    print(f"文件: {os.path.basename(path)}")
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True, max_row=4))
        print(f"\n  [{name}] 前4行:")
        for r in rows:
            print(f"    {[str(v)[:20] if v is not None else None for v in r[:15]]}")
    wb.close()


# ── 读取干员满级数据（终末地干员数据 sheet，等级=90）─────────────────────
def extract_characters():
    path = os.path.join(DATA_DIR, "终末地干员数据.xlsm")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["终末地干员数据"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    print("干员数据列头:", headers)

    chars = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(len(headers))}
        chars.append(d)

    print(f"共 {len(chars)} 条干员数据")
    if chars:
        print("第一条:", {k: v for k, v in list(chars[0].items())[:10]})
    return chars


# ── 读取装备数据 ─────────────────────────────────────────────────────────
def extract_equipment():
    path = os.path.join(DATA_DIR, "终末地装备数据.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("\n装备文件 sheets:", wb.sheetnames)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True, max_row=5))
        print(f"\n  [{name}] 前5行:")
        for r in rows:
            print(f"    {[str(v)[:25] if v is not None else None for v in r[:12]]}")
        result[name] = read_sheet_as_dicts(ws)
    wb.close()
    return result


chars = extract_characters()
equip = extract_equipment()

# 保存供下一步使用
with open(os.path.join(DATA_DIR, "extracted_characters.json"), "w", encoding="utf-8") as f:
    json.dump(chars, f, ensure_ascii=False, indent=2)

with open(os.path.join(DATA_DIR, "extracted_equipment.json"), "w", encoding="utf-8") as f:
    json.dump(equip, f, ensure_ascii=False, indent=2)

print("\n已保存 extracted_characters.json 和 extracted_equipment.json")
